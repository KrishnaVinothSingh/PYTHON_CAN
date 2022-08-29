import platform
import ctypes
import sys
import xlrd
import xlwt
import xlsxwriter
from openpyxl import load_workbook

ostype = ctypes.sizeof(ctypes.c_voidp) * 8
if (platform.system()=='Windows') or (platform.system()[0:6]=='CYGWIN') :
    # WINDOWS
    t32api = ctypes.CDLL("D:/Programme/Trace32/TriCore/ICD/19_09_01\demo/api/capi/dll/t32api64.dll" if ostype==64 else "D:/Programme/Trace32/TriCore/ICD/19_09_01\demo/api/capi/dll/t32api.dll")
    print("Pass")
elif platform.system()=='Darwin' :
    # Mac OS X
    t32api = ctypes.CDLL("./t32api.dylib")
else :
    # Linux
    t32api = ctypes.CDLL("./t32api64.so" if ostype==64 else  "./t32api.so")


# Declare UDP/IP socket of the TRACE32 instance to access
t32api.T32_Config(b"NODE=",b"localhost")
t32api.T32_Config(b"PORT=",b"20000")
print("pass2")

# Connect to TRACE32
error = t32api.T32_Init()
print("pass3")
if error != 0 :
   sys.exit("Can't connect to TRACE32!")

# Select to debugger component of TRACE32 (B:: prompt)
t32api.T32_Attach(1)



# Start program
t32api.T32_Go()
print("pass2")

#write the values to excelsheet
wb = load_workbook('D:/Krishna/Trace_demo1.xlsx')  # Work Book
sheet = wb.active
#Read the Variable address from excel sheet 
workbook = xlrd.open_workbook('D:/Krishna/Trace_demo1.xlsx')
worksheet = workbook.sheet_by_name('Sheet1')
num_rows = worksheet.nrows - 1
print(num_rows)
curr_row = 0
while curr_row < num_rows:
        curr_row += 1
        row = worksheet.row(curr_row)
        print(curr_row)
        row=(row[0].value)
        row=bytes(str(row).encode("utf-8"))
        print(row)
        print(type(row))
        #row=int(row,base=16)
        #print (type(row))

# Read variable
        vname = row
        vvalue = ctypes.c_int32(0)
        vvalueh = ctypes.c_int32(0)
        rc = t32api.T32_ReadVariableValue(vname,ctypes.byref(vvalue),ctypes.byref(vvalueh))
        if rc == 0 :
            print (rc)
            print("flags[3] = " + str(vvalue.value))
            data32 = str(vvalue.value)            
            sheet.cell(row=(curr_row+1),column=3).value=data32
            wb.save('D:/Krishna/Trace_demo1.xlsx')

        else:
            print("read failed")
        
        







